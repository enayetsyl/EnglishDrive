#!/usr/bin/env python3
"""
English Skill-Building Drive — programmatic audit suite (Run Book §6.5 / §6.8 / §6.11).

Usage:
    python3 run_all.py <manifest.json> [--file2 <path.xlsx>] [--report-dir <dir>]

The manifest is a machine-readable extraction of every graded answer set, item text,
stated mark total, and dictation list from the draft sheets. Schema:
audits/scripts/README_manifest.md

Exit code 0 = all gates PASS (human-review flags may still be present).
Exit code 1 = one or more gates FAIL.

All string comparisons are punctuation- and case-normalised on BOTH sides
(Starter Template §B: a comma once defeated an audit).
"""

import json
import re
import sys
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path

# ----------------------------------------------------------------------------- helpers

def norm(s: str) -> str:
    """Normalise for comparison: NFC, lowercase, strip all punctuation, collapse ws."""
    s = unicodedata.normalize("NFC", str(s))
    s = s.lower()
    s = re.sub(r"[^\w\s\u0980-\u09FF]", " ", s)   # keep word chars + Bangla range
    s = re.sub(r"\s+", " ", s).strip()
    return s


def graded_sheets(m):
    """Sheets that belong to THIS block's graded surface.

    A sheet may carry "audit_scope": "pt_overlap_only" to be visible ONLY to
    gate_pt_zero_overlap() — used to bring another block's already-validated
    worksheets into the PT overlap comparison without re-running the
    de-patterning / marks / held-word gates over them.

    Authorised for two shapes:
      - PD-032 — a two-week `a`/`b` split block whose b-half PT grades a-half grammar.
      - PD-034 — a §6.7 paired-week recovery, where one week's PT was built but not
        administered and its assessment is carried into a combined PT for the
        following week (PD-029: `C4B05-PT` → `C4B0506-PT`). The carried PT must be
        verified against BOTH blocks' worksheets and the unadministered PT itself,
        which is exactly this flag's comparison surface.

    Additive: a sheet with no "audit_scope" key is graded, so a manifest that
    declares none behaves exactly as before.
    """
    return [s for s in m["sheets"] if s.get("audit_scope") != "pt_overlap_only"]


def all_items(sheet):
    for part in sheet.get("parts", []):
        for it in part.get("items", []):
            yield part.get("name", "?"), it


# --------------------------------------------------------------- PD-054 surfaces

def student_facing_surfaces(sheet):
    """Yield (location, kind, text) for student-facing text that is NOT item text.

    PD-054. Before this, every gate read `item["text"]`, `item["trigger"]` or
    `part["options"]` and nothing else, so three printed surfaces were invisible:

      kind="clue"          item-level gloss, e.g. "(সঙ্গে)" — answer-bearing support
      kind="instructions"  part-level rubric line, e.g. "Fill in the blanks…"
      kind="boxes"         sheet-level word banks / support + fading hint boxes

    Clue glosses were previously covered only by accident: `(সঙ্গে)` sat inside
    the C4B06 `text` strings because the extractor happened to leave the
    parenthetical inline. An extractor that stripped it — the natural choice,
    since it is not part of the sentence — would have blinded every gate to it.
    The fields make that coverage declared instead of incidental.

    All three fields are optional; a manifest declaring none behaves exactly as
    before.
    """
    for b in sheet.get("boxes", []) or []:
        if isinstance(b, dict):
            label = b.get("name", "box")
            content = " ".join(str(w) for w in b.get("words", []) or []) \
                      + " " + str(b.get("text", "") or "")
        else:
            label, content = "box", str(b)
        if content.strip():
            yield (f"{sheet['name']} {label}", "boxes", content.strip())
    for part in sheet.get("parts", []):
        pname = part.get("name", "?")
        instr = part.get("instructions")
        if instr:
            yield (f"{sheet['name']} Part {pname} instructions", "instructions", str(instr))
        for i, it in enumerate(part.get("items", []), 1):
            clue = it.get("clue")
            if clue:
                yield (f"{sheet['name']} Part {pname} #{i} clue", "clue", str(clue))


# English function words: never "content" words, so never held-word targets.
CLUE_FUNCTION_WORDS = {
    "a", "an", "the", "and", "or", "but", "of", "to", "in", "on", "at", "by",
    "for", "with", "from", "into", "onto", "up", "down", "out", "over", "under",
    "is", "are", "was", "were", "be", "been", "am", "do", "does", "did", "has",
    "have", "had", "will", "would", "can", "could", "shall", "should", "may",
    "might", "must", "not", "no", "yes", "this", "that", "these", "those",
    "here", "there", "it", "its", "he", "she", "they", "we", "you", "i", "him",
    "her", "them", "us", "me", "my", "your", "his", "their", "our", "who",
    "what", "which", "where", "when", "why", "how", "if", "then", "than", "as",
    "so", "too", "very", "one", "two", "three", "four", "five",
}

# Charter §H.5 house roster — sanctioned by name, never pool vocabulary.
ROSTER_NAMES = {
    "yusuf", "abdullah", "nusair", "abdur", "rahim",
    "aisha", "raima", "maryam", "fatima", "porshi", "rabab", "jesmin",
}


def _is_english_token(tok):
    """True for a pure a–z token. Bangla glosses return False and are exempt."""
    return bool(tok) and all("a" <= c <= "z" for c in tok)


def sheet_answer_sets(sheet):
    """Yield (part_name, [answers]) for every part that has an answer sequence."""
    for part in sheet.get("parts", []):
        answers = part.get("answers")
        if answers is None:
            answers = [it.get("answer") for it in part.get("items", [])
                       if it.get("answer") is not None]
        if answers:
            yield part.get("name", "?"), [str(a) for a in answers]


# ----------------------------------------------------------------------------- gates

def gate_depattern(m):
    """Max run ≤2 and no strict alternation, per graded answer set."""
    fails, checked = [], 0
    for sheet in graded_sheets(m):
        for pname, answers in sheet_answer_sets(sheet):
            checked += 1
            a = [norm(x) for x in answers]
            # max run
            run, best = 1, 1
            for i in range(1, len(a)):
                run = run + 1 if a[i] == a[i - 1] else 1
                best = max(best, run)
            if best > 2:
                fails.append(f"{sheet['name']} Part {pname}: run of {best}")
            # strict alternation: whole set flips between exactly two values, len ≥ 4
            if len(a) >= 4 and len(set(a)) == 2:
                if all(a[i] != a[i - 1] for i in range(1, len(a))):
                    fails.append(f"{sheet['name']} Part {pname}: strict alternation")
    return ("De-patterning", not fails, f"{checked} answer sets checked", fails)


def gate_pair_overlap(m):
    """CW↔HW positional answer overlap ≤35%; identical item texts ≤2 per day."""
    fails, notes = [], []
    by_name = {s["name"]: s for s in graded_sheets(m)}
    for sheet in graded_sheets(m):
        pair = sheet.get("pair")
        if not pair or pair not in by_name:
            continue
        other = by_name[pair]
        # positional answer overlap, part-aligned by order
        sa = [x for _, xs in sheet_answer_sets(sheet) for x in xs]
        oa = [x for _, xs in sheet_answer_sets(other) for x in xs]
        n = min(len(sa), len(oa))
        if n:
            same = sum(1 for i in range(n) if norm(sa[i]) == norm(oa[i]))
            pct = 100.0 * same / n
            notes.append(f"{sheet['name']}↔{pair}: positional {pct:.0f}%")
            if pct > 35.0:
                fails.append(f"{sheet['name']}↔{pair}: positional overlap {pct:.0f}% (> 35%)")
        # identical item texts
        st = {norm(it.get("text", "")) for _, it in all_items(sheet) if it.get("text")}
        ot = [norm(it.get("text", "")) for _, it in all_items(other) if it.get("text")]
        ident = sum(1 for t in ot if t in st)
        if ident > 2:
            fails.append(f"{sheet['name']}↔{pair}: {ident} identical item texts (> 2)")
    note = ("; ".join(notes) or "no pairs declared") + \
        " (superseded by PD-036 zero-repeat, kept as backstop)"
    return ("CW↔HW overlap", not fails, note, fails)


def gate_pt_zero_overlap(m):
    """Zero PT item texts identical (normalised) to any worksheet item."""
    fails = []
    pts = [s for s in m["sheets"] if s.get("type") == "pt"]
    wss = [s for s in m["sheets"] if s.get("type") != "pt"]
    ws_texts = {norm(it.get("text", "")): s["name"]
                for s in wss for _, it in all_items(s) if it.get("text")}
    for pt in pts:
        for pname, it in all_items(pt):
            t = norm(it.get("text", ""))
            if t and t in ws_texts:
                fails.append(f"{pt['name']} Part {pname}: item duplicates {ws_texts[t]}: "
                             f"\"{it.get('text','')[:60]}\"")
    return ("PT zero-overlap", not fails,
            f"{sum(1 for s in pts for _ in all_items(s))} PT items vs "
            f"{len(ws_texts)} worksheet items", fails)


def gate_within_sheet_dupes(m):
    fails = []
    for sheet in graded_sheets(m):
        texts = [norm(it.get("text", "")) for _, it in all_items(sheet) if it.get("text")]
        for t, c in Counter(texts).items():
            if c > 1:
                fails.append(f"{sheet['name']}: duplicate item ×{c}: \"{t[:60]}\"")
    return ("Within-sheet duplicates", not fails, "all sheets scanned", fails)


def gate_option_list(m):
    """Every keyed answer in a part must appear in that part's printed option list.

    CR-008 / PD-037: six sheets historically keyed an answer (Adverb) that the
    printed instruction options excluded — the one defect a pupil meets directly.
    Additive: only parts that declare an "options" field are checked, so existing
    manifests without the field behave exactly as before (the note says vacuous).
    """
    fails, checked = [], 0
    for sheet in graded_sheets(m):
        for part in sheet.get("parts", []):
            opts = part.get("options")
            if not opts:
                continue
            checked += 1
            nopts = {norm(o) for o in opts}
            for it in part.get("items", []):
                a = it.get("answer")
                if a is not None and norm(a) not in nopts:
                    fails.append(f"{sheet['name']} Part {part.get('name', '?')}: keyed "
                                 f"answer '{a}' missing from printed options {opts}")
    note = (f"{checked} option-listed parts checked" if checked else
            "no parts declare 'options' — gate vacuous; declare on identify parts")
    return ("Option-list completeness", not fails, note, fails)


def gate_hw_key(m):
    """No part's answer sequence may be positionally identical to the same-named
    part of its paired sheet — the HW key must not be transcribable from the CW.

    CR-006 / PD-037: the whole-sheet ≤35% positional gate passed C4B06 while five
    individual HW parts carried keys identical to their CW counterparts. Compares
    same-named parts of a declared pair, sequences of length ≥3 and equal length.
    """
    fails, checked = [], 0
    by_name = {s["name"]: s for s in graded_sheets(m)}
    seen = set()
    for sheet in graded_sheets(m):
        pair = sheet.get("pair")
        if not pair or pair not in by_name:
            continue
        key = tuple(sorted((sheet["name"], pair)))
        if key in seen:
            continue
        seen.add(key)
        other = by_name[pair]
        oparts = {p: a for p, a in sheet_answer_sets(other)}
        for pname, answers in sheet_answer_sets(sheet):
            oa = oparts.get(pname)
            if oa is None or len(answers) < 3 or len(answers) != len(oa):
                continue
            checked += 1
            if [norm(x) for x in answers] == [norm(x) for x in oa]:
                fails.append(f"{sheet['name']}↔{other['name']} Part {pname}: identical "
                             f"answer sequence ({' · '.join(answers)})")
    return ("HW key transcribability", not fails, f"{checked} paired parts compared", fails)


# PD-036: max identical carrier sentences tolerated across a block's sheets.
# 0 = every normalised item text unique across all graded sheets (the C4B06
# promotion standard: 199 sentences, zero repeats). Raise to loosen.
CROSS_SHEET_MAX_REPEATS = 0

def gate_cross_sheet_repetition(m):
    """No item text may appear on more than one of the block's graded sheets.

    CR-009 / PD-036: C4B06 reached review with 38 sentences repeated over 78
    placements (CW-4 was 16/20 recycled). Threshold: a text on N sheets fails
    when (N - 1) > CROSS_SHEET_MAX_REPEATS. Strictly tighter than the CW↔HW
    "≤2 identical items per day" allowance, which can never bind while this
    gate holds at 0.
    """
    where = {}
    for sheet in graded_sheets(m):
        for _pname, it in all_items(sheet):
            t = norm(it.get("text", ""))
            if t:
                where.setdefault(t, set()).add(sheet["name"])
    fails = [f"\"{t[:60]}\" on {len(s)} sheets: {', '.join(sorted(s))}"
             for t, s in sorted(where.items())
             if (len(s) - 1) > CROSS_SHEET_MAX_REPEATS]
    return ("Cross-sheet repetition (PD-036)", not fails,
            f"{len(where)} distinct texts across {len(graded_sheets(m))} sheets; "
            f"threshold {CROSS_SHEET_MAX_REPEATS}", fails)


def gate_rehearsal_disjoint(m):
    st = {norm(w) for w in m.get("pt_self_try_words", [])}
    demo = {norm(w) for w in m.get("demo_box_words", [])}
    shared = sorted(st & demo)
    if not st and not demo:
        return ("Rehearsal/graded disjoint", True, "not applicable (no boxes declared)", [])
    return ("Rehearsal/graded disjoint", not shared,
            f"{len(st)} self-try vs {len(demo)} demo words",
            [f"shared word: {w}" for w in shared])


def gate_marks(m):
    fails, notes = [], []
    for sheet in graded_sheets(m):
        stated = sheet.get("stated_total")
        per = []
        for part in sheet.get("parts", []):
            pm = part.get("marks")
            if pm is not None:
                per.append(float(pm))
            else:
                per.extend(float(it.get("marks", 1)) for it in part.get("items", []))
        comp = sum(per)
        if stated is None:
            notes.append(f"{sheet['name']}: no stated total (computed {comp:g})")
            continue
        notes.append(f"{sheet['name']}: {comp:g}/{stated:g}")
        if abs(comp - float(stated)) > 1e-9:
            fails.append(f"{sheet['name']}: computed {comp:g} ≠ stated {stated:g}")
    return ("Mark totals", not fails, "; ".join(notes), fails)


SACRED = {"allah", "আল্লাহ", "quran", "qur an", "কুরআন", "কোরআন"}

def gate_sacred(m):
    fails = []
    for sheet in graded_sheets(m):
        for pname, it in all_items(sheet):
            for field in ("answer", "trigger"):
                v = it.get(field)
                if v and norm(v) in SACRED:
                    fails.append(f"{sheet['name']} Part {pname}: sacred word as graded "
                                 f"{field}: {v}")
    for w in m.get("dictation", []):
        if norm(w) in SACRED:
            fails.append(f"dictation: sacred word graded: {w}")
    # PD-054: the three student-facing surfaces, split by Charter §H.3, which bars a
    # sacred word as a graded CLASSIFICATION TARGET but permits teacher prose.
    #   clue  → FAIL. A gloss is answer-bearing support, functionally part of the item.
    #   boxes → FAIL. A word bank is a selectable option list, i.e. a graded target.
    #   instructions → FLAG. Rubric wording is the teacher prose §H.3 permits, but it
    #                  is surfaced for the human read rather than passed silently.
    flags = []
    for sheet in graded_sheets(m):
        for where, kind, text in student_facing_surfaces(sheet):
            hits = sorted({t for t in norm(text).split() if t in SACRED})
            if not hits:
                continue
            msg = f"{where}: sacred word in {kind}: {', '.join(hits)}"
            (flags if kind == "instructions" else fails).append(msg)
    note = "graded targets + dictation + clue/instructions/boxes scanned (PD-054)"
    if flags:
        note += f" — {len(flags)} instruction-line hit(s) flagged for human read"
    return ("Sacred-word guard", not fails, note, fails + flags)


VALUES_LEXICON = {
    "music", "song", "sing", "dance", "guitar", "drum", "piano", "flute", "violin",
    "band", "concert", "christmas", "halloween", "easter", "diwali", "puja", "holi",
    "valentine", "গান", "নাচ", "পূজা", "বড়দিন",
}

# Irregular inflections that stem matching cannot reach from the lexicon entry.
VALUES_IRREGULAR = {
    "sang": "sing", "sung": "sing", "sunge": "sing",
    "danced": "dance", "dancing": "dance",
    "drummed": "drum", "drumming": "drum",
}

def _values_hits(text):
    """Match lexicon entries across inflections (T2).

    The lexicon stores base forms, but student-facing text carries inflected ones
    (sings / sang / singing / songs). Whole-token equality missed five of six
    singing items in the C4B06 review, so matching now covers:
      - exact token            sing
      - regular suffixes       sings, singing, songs, dances, dancing
      - irregular forms        sang, sung   (VALUES_IRREGULAR)
    Prefix matching is length-guarded (base ≥4 chars) so short entries cannot
    over-match; 'band' will not fire on 'bandage' because the suffix must be one
    of the inflection endings, not arbitrary text.
    """
    SUFFIXES = ("", "s", "es", "d", "ed", "ing", "er", "ers")
    hits = set()
    for tok in norm(text).split():
        if tok in VALUES_LEXICON:
            hits.add(tok)
            continue
        if tok in VALUES_IRREGULAR:
            hits.add(VALUES_IRREGULAR[tok])
            continue
        for base in VALUES_LEXICON:
            if len(base) < 4:
                continue
            for suf in SUFFIXES:
                if tok == base + suf:
                    hits.add(base)
                    break
            # verbs ending -e drop it before -ing/-ed  (dance -> dancing)
            if base.endswith("e") and tok in (base[:-1] + "ing", base[:-1] + "ed"):
                hits.add(base)
    return hits


def gate_values_lexicon(m):
    flags = []
    for sheet in graded_sheets(m):
        for pname, it in all_items(sheet):
            hit = _values_hits(it.get("text", ""))
            if hit:
                flags.append(f"{sheet['name']} Part {pname}: {sorted(hit)} in "
                             f"\"{it.get('text','')[:60]}\"")
        # PD-054: clue glosses, part instruction lines and sheet boxes are
        # student-facing print and are screened on the same footing as item text.
        for where, kind, text in student_facing_surfaces(sheet):
            hit = _values_hits(text)
            if hit:
                flags.append(f"{where} ({kind}): {sorted(hit)} in \"{text[:60]}\"")
    # This gate FLAGS for human review; lexicon hits are treated as failures until ruled.
    return ("Values lexicon screen", not flags,
            "item texts + clue/instructions/boxes scanned (incl. inflections; PD-054)",
            flags)


def load_file2_words(path):
    """Heuristic loader: find a word column and (if present) a release-week column.

    Sheets with an EXACT 'word' header are preferred; fuzzy matches ("words this
    week") are used only if no sheet in the workbook has an exact word column.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        exact = next((i for i, h in enumerate(header)
                      if h in ("word", "words", "english", "english word")), None)
        fuzzy = next((i for i, h in enumerate(header) if "word" in h), None)
        kcol = next((i for i, h in enumerate(header)
                     if any(k in h for k in ("week", "release", "batch"))), None)
        if exact is not None or fuzzy is not None:
            sheets.append((exact, fuzzy, kcol, rows))
    use_exact = any(e is not None for e, _, _, _ in sheets)
    words = {}
    for exact, fuzzy, kcol, rows in sheets:
        wcol = exact if use_exact else fuzzy
        if wcol is None:
            continue
        for r in rows[1:]:
            if r[wcol] is None:
                continue
            w = norm(r[wcol])
            wk = None
            if kcol is not None and r[kcol] is not None:
                mnum = re.search(r"\d+", str(r[kcol]))
                wk = int(mnum.group()) if mnum else None
            if w and (w not in words or (wk is not None and
                       (words[w] is None or wk < words[w]))):
                words[w] = wk
    return words


def gate_heldword(m, file2_path):
    if not file2_path:
        return ("Held-word / exemplar", False,
                "SKIPPED — no --file2 given (gate cannot pass without it)",
                ["provide --file2 <pool.xlsx>"])
    try:
        pool = load_file2_words(file2_path)
    except Exception as e:  # noqa: BLE001
        return ("Held-word / exemplar", False, f"File 2 load error: {e}", [str(e)])
    if not pool:
        return ("Held-word / exemplar", False,
                "File 2 loaded but no word column recognised — check README_manifest.md",
                ["unrecognised File 2 layout"])
    exemplars = {norm(w) for w in m.get("exemplars", [])}
    # PD-012 block-local teaching set (T1): taught in-block, gradeable in-block,
    # NOT in File 2, NOT held downstream, and NEVER a dictation/spelling item.
    # Declared separately from exemplars because the two are different instruments
    # (PD-009 exemplars are rule-demonstrators and are not vocabulary at all), and
    # because the dictation prohibition below only applies to block-local words.
    block_local = {norm(w) for w in m.get("block_local", [])}
    overlap = exemplars & block_local
    build_week = m.get("build_week")
    fails, checked = [], 0
    if overlap:
        fails.append("declared as BOTH exemplar and block_local: "
                     + ", ".join(sorted(overlap)))
    # PD-012: block-local words are never dictation or spelling items.
    for w in m.get("dictation", []):
        if norm(w) in block_local:
            fails.append(f"dictation: '{w}' is block-local (PD-012 bars it from dictation)")
    targets = []
    for sheet in graded_sheets(m):
        for pname, it in all_items(sheet):
            trig = it.get("trigger")
            if trig:
                targets.append((f"{sheet['name']} {pname}", trig))
    for w in m.get("dictation", []):
        targets.append(("dictation", w))
    for where, w in targets:
        checked += 1
        nw = norm(w)
        if nw in exemplars or nw in block_local:
            continue
        if nw not in pool:
            fails.append(f"{where}: '{w}' not in File 2 pool and not a declared "
                         f"exemplar or block-local word")
        elif build_week is not None and pool[nw] is not None and pool[nw] > build_week:
            fails.append(f"{where}: '{w}' released week {pool[nw]} > build week {build_week}")
    # PD-054: clue glosses carry the held-word obligation; instructions/boxes do not.
    #
    # A clue that carries the answer is functionally part of the item, which is why
    # `(together)` — a W7 word — could print on a W6 sheet, key five graded items,
    # and leave every gate green: the held-word gate reads `trigger` alone.
    # Instruction and rubric language is exempt: "Fill in the blanks" is teacher
    # register, not vocabulary, and holding it to the pool would be meaningless.
    # Bangla glosses are exempt by nature (CR-011 makes Bangla the required route) —
    # they are not pool words at all — but they are still values/sacred screened.
    clue_checked = 0
    for sheet in graded_sheets(m):
        for where, kind, text in student_facing_surfaces(sheet):
            if kind != "clue":
                continue
            for tok in norm(text).split():
                if not _is_english_token(tok):
                    continue          # Bangla / mixed script — exempt (CR-011)
                if tok in CLUE_FUNCTION_WORDS or tok in ROSTER_NAMES:
                    continue
                clue_checked += 1
                if tok in exemplars or tok in block_local:
                    continue
                if tok in pool and not (build_week is not None
                                        and pool[tok] is not None
                                        and pool[tok] > build_week):
                    continue
                if tok not in pool:
                    fails.append(f"{where}: clue word '{tok}' not in File 2 pool and "
                                 f"not a declared exemplar or block-local word (PD-054)")
                else:
                    fails.append(f"{where}: clue word '{tok}' released week "
                                 f"{pool[tok]} > build week {build_week} (PD-054)")
    note = (f"{checked} graded targets checked against {len(pool)} pool words"
            f" + {len(exemplars)} exemplars + {len(block_local)} block-local"
            f"; {clue_checked} English clue word(s) checked (PD-054)")
    if not targets:
        note += " — WARNING: manifest declares no 'trigger' fields; gate is vacuous"
    return ("Held-word / exemplar / block-local", not fails, note, fails)


def gate_one_defensible(m):
    """Cannot be fully automated — surface candidates where one item text maps to
    multiple distinct answers across the block, for HUMAN review."""
    seen = {}
    flags = []
    for sheet in graded_sheets(m):
        for pname, it in all_items(sheet):
            t, a = norm(it.get("text", "")), norm(it.get("answer", "") or "")
            if not t or not a:
                continue
            if t in seen and seen[t][0] != a:
                flags.append(f"same item, different answers: \"{t[:50]}\" → "
                             f"{seen[t][0]} ({seen[t][1]}) vs {a} ({sheet['name']} {pname})")
            seen.setdefault(t, (a, f"{sheet['name']} {pname}"))
    # informational: never blocks on its own, but is printed for the human pass
    return ("One-defensible-answer (human)", True,
            f"{len(flags)} candidates for human review" if flags else
            "no conflicting-answer candidates; full check remains a human read", flags)


# ----------------------------------------------------------------------------- main

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    manifest_path = Path(sys.argv[1])
    file2 = None
    report_dir = Path(__file__).resolve().parent.parent / "reports"
    args = sys.argv[2:]
    for i, a in enumerate(args):
        if a == "--file2" and i + 1 < len(args):
            file2 = args[i + 1]
        if a == "--report-dir" and i + 1 < len(args):
            report_dir = Path(args[i + 1])

    m = json.loads(manifest_path.read_text(encoding="utf-8"))
    block = m.get("block_id", manifest_path.stem)

    gates = [
        gate_depattern(m),
        gate_pair_overlap(m),
        gate_pt_zero_overlap(m),
        gate_within_sheet_dupes(m),
        gate_option_list(m),
        gate_hw_key(m),
        gate_cross_sheet_repetition(m),
        gate_rehearsal_disjoint(m),
        gate_marks(m),
        gate_sacred(m),
        gate_values_lexicon(m),
        gate_heldword(m, file2),
        gate_one_defensible(m),
    ]

    lines = [f"AUDIT REPORT — {block} — {date.today().isoformat()}",
             f"manifest: {manifest_path}", f"file2: {file2 or 'NOT PROVIDED'}", "-" * 72]
    ok = True
    for name, passed, note, detail in gates:
        status = "PASS" if passed else "FAIL"
        if not passed:
            ok = False
        lines.append(f"[{status}] {name:32s} {note}")
        for d in detail:
            lines.append(f"        - {d}")
    lines.append("-" * 72)
    lines.append("RESULT: ALL GATES PASS" if ok else "RESULT: FAILURES PRESENT — do not finalize")
    lines.append("Reminder: the naturalness gate and full one-defensible-answer check are HUMAN.")
    out = "\n".join(lines)
    print(out)

    report_dir.mkdir(parents=True, exist_ok=True)
    rp = report_dir / f"{block}_audit_{date.today().isoformat()}.txt"
    rp.write_text(out, encoding="utf-8")
    print(f"\nreport saved: {rp}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
